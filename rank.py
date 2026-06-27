#!/usr/bin/env python3
"""
Redrob Hackathon — Candidate Ranker v2
Senior AI Engineer (Founding Team) at Redrob AI

Improvements over v1 (per RankPy_Improvements_Guide):
 1.  Reduce double-counting: retrieval/ranking split out of career, not also double-weighted in title
 2.  Career recency weighting: exponential decay so recent roles count more
 3.  Keyword stuffing detection: unique AI concept count, cap repeated hits
 4.  Production AI signals: boost for production/deployment/inference/serving/latency/monitoring
 5.  Career progression: reward Engineer→Senior→Staff; penalize regressions
 6.  Better consulting detection: company + industry + description together
 7.  Buzzword profile detection: many AI buzzwords but no implementation evidence
 8.  Cross-validate skills: trust boost if skill also appears in job descriptions/titles
 9.  Skill synergy bonus: Python + Retrieval + Embeddings + VectorDB + LLM combo
10.  Smooth notice period: continuous scoring (no hard buckets)
11.  Better score normalization: z-score + sigmoid on final composite
12.  Penalize shallow skill lists: large list, low proficiency/duration/endorsements
13.  End-to-end ML lifecycle: reward training + eval + deployment + monitoring coverage
14.  Dynamic title compensation: boost generic titles with verified AI career evidence
15.  Small company reputation bonus: strong AI-product companies (verified)
16.  Improve retrieval weight: highest skill but not dominating entire score
17.  Project impact signals: latency, recall, NDCG, MRR, QPS, scale metrics
18.  Open source contribution: GitHub activity + OSS mentions
19.  Leadership: mentoring, technical leadership for senior roles
20.  Explainability: per-component scores stored and returned

Targeted fixes (production review):
 A.  Company reputation bonus removed: STRONG_AI_COMPANIES emptied. Ranking is
     now based entirely on demonstrated skills, career evidence, and behavioral
     signals — employer name is not evidence of skill.
 B.  Career recency anchored on end_date (or TODAY for current roles) instead
     of start_date. A long-running current role is as recent as a brand-new
     one; the old anchor incorrectly penalised candidates mid-tenure.
 C.  generate_reasoning() now uses the same precompiled regex patterns
     (RETRIEVAL_RANKING_PATTERNS, AI_KEYWORD_PATTERNS, MANDATORY_SKILL_PATTERNS,
     OPTIONAL_SKILL_PATTERNS) as the scoring functions, eliminating false
     positives from plain substring matching (e.g. "rag" in "storage") and
     ensuring reasoning is fully consistent with how candidates were ranked.
 D.  Output format changed from CSV to XLSX (submission portal only accepts
     "excel, spreadsheet" files — see main()).
"""

import json
import re
import math
from datetime import datetime, date
from pathlib import Path

# ============================================================
# CONSTANTS & SIGNAL SETS
# ============================================================

TODAY = date(2026, 6, 25)

CONSULTING_FIRMS = {
    'tcs', 'infosys', 'wipro', 'accenture', 'cognizant', 'capgemini',
    'hcl', 'mphasis', 'hexaware', 'tech mahindra', 'l&t technology',
    'ltimindtree', 'mindtree', 'persistent systems', 'niit technologies',
}
CONSULTING_INDUSTRIES = {'it services', 'consulting', 'outsourcing', 'bpo', 'staffing'}

STRONG_AI_COMPANIES: set = set()

MANDATORY_SKILLS = {
    'sentence-transformers', 'sentence transformers', 'embeddings', 'vector embeddings',
    'faiss', 'pinecone', 'weaviate', 'qdrant', 'milvus', 'opensearch',
    'elasticsearch', 'dense retrieval', 'hybrid search', 'vector search',
    'vector database', 'vector db', 'rag', 'retrieval augmented',
    'bge', 'e5 embeddings', 'learning to rank', 'ndcg', 'mrr', 'map@',
    'ranking', 'information retrieval', 'recommendation systems',
    'recommender', 'search ranking', 'reranking', 'python',
}

OPTIONAL_SKILLS = {
    'pytorch', 'tensorflow', 'transformers', 'huggingface', 'hugging face',
    'scikit-learn', 'sklearn', 'xgboost', 'lightgbm',
    'lora', 'qlora', 'peft', 'fine-tuning llms', 'fine-tuning', 'llm',
    'llms', 'large language models', 'gpt', 'bert', 'nlp',
}

SYNERGY_GROUPS = [
    {'python', 'embeddings', 'retrieval', 'vector', 'llm'},
    {'faiss', 'elasticsearch', 'retrieval', 'ranking', 'python'},
    {'rag', 'llm', 'embeddings', 'vector database', 'python'},
    {'fine-tuning', 'transformers', 'embeddings', 'ranking', 'python'},
]

NEGATIVE_SKILLS = {
    'marketing', 'photoshop', 'illustrator', 'figma', 'sap', 'oracle',
    'salesforce', 'crm', 'powerpoint', '.net', 'c#', 'php',
    'ruby', 'swift', 'react native', 'flutter', 'unity', 'unreal',
    'autocad', 'solidworks',
}

PREFERRED_LOCATIONS_TOP = {'pune', 'noida'}
PREFERRED_LOCATIONS = {
    'pune', 'noida', 'bangalore', 'bengaluru', 'hyderabad', 'mumbai',
    'delhi', 'gurgaon', 'gurugram', 'chennai', 'kolkata', 'ahmedabad', 'india',
}

GOOD_TITLE_PATTERNS = [
    r'\bai\b', r'\bml\b', r'machine learning', r'deep learning',
    r'nlp\b', r'natural language', r'search engineer', r'ranking engineer',
    r'retrieval', r'recommendation', r'applied (science|ml|ai|research)',
    r'data scientist', r'research engineer', r'llm engineer',
    r'generative ai', r'embedding', r'senior engineer',
]

BAD_TITLE_PATTERNS = [
    r'\bmarketing\b', r'\bhr\b', r'human resources', r'\bfinance\b',
    r'\baccountant\b', r'\bsales\b', r'\bdesigner\b', r'\bux\b',
    r'\bui designer\b', r'product manager', r'project manager',
    r'\brecruiter\b', r'\bqa\b', r'quality assurance',
    r'mobile developer', r'android', r'ios developer',
    r'\bphp\b', r'\bruby\b', r'\bsap\b', r'\boracle\b',
]

TITLE_TIERS = [
    (1.00, [r'\bml engineer', r'machine learning engineer', r'\bai engineer',
            r'applied (scientist|science)', r'research engineer',
            r'llm engineer', r'\bnlp engineer', r'search engineer',
            r'ranking engineer', r'retrieval engineer', r'recommendation(s)? engineer']),
    (0.65, [r'backend.*(ml|ai|machine learning)', r'(ml|ai|machine learning).*backend',
            r'data scientist', r'deep learning engineer', r'computer vision engineer',
            r'\bmlops\b']),
    (0.45, [r'data engineer', r'analytics engineer', r'\bml ops\b']),
    (0.35, [r'\bsoftware engineer\b', r'\bsde\b', r'backend engineer',
            r'full[\s-]?stack engineer', r'platform engineer']),
    (0.18, [r'\bconsultant\b', r'\btechnical consultant\b', r'solutions? consultant']),
]

SENIORITY_ORDER = {
    'intern': 0, 'trainee': 0, 'junior': 1, 'jr': 1, 'associate': 2,
    'engineer': 3, 'developer': 3, 'scientist': 3, 'analyst': 3,
    'senior': 4, 'lead': 4, 'sr': 4,
    'staff': 5, 'principal': 6,
    'director': 7, 'head': 7, 'vp': 8, 'chief': 9, 'cto': 9,
}

MIN_YOE_FOR_LEVEL = {0: 0, 1: 0, 2: 1, 3: 1, 4: 4, 5: 6, 6: 8, 7: 9, 8: 11, 9: 12}

def seniority_plausibility_factor(title, yoe):
    level = _seniority_level(title)
    min_yoe = MIN_YOE_FOR_LEVEL.get(level, 0)
    if min_yoe == 0 or yoe >= min_yoe:
        return 1.0
    ratio = yoe / min_yoe
    return max(0.55, 0.55 + 0.45 * ratio)

PRODUCTION_KW = [
    'production', 'deployed', 'deployment', 'inference', 'serving',
    'latency', 'monitoring', 'pipeline', 'real-time', 'online',
    'a/b test', 'ab test', 'rollout', 'canary', 'sla', 'throughput',
]
IMPACT_KW = [
    'latency reduction', 'recall@', 'precision@', 'ndcg', 'mrr',
    'qps', 'queries per second', 'scale', 'billion', 'million',
    'reduced', 'improved', 'increased', 'decreased', '%', 'x faster',
    'ms p99', 'p95', 'p50',
]
OSS_KW = ['open.source', 'github', 'open source', 'contributed', 'maintainer', 'contributor']
LEADERSHIP_KW = ['mentor', 'lead', 'managed', 'built the team', 'hired', 'technical lead',
                 'tech lead', 'ownership', 'principal', 'architect']
LIFECYCLE_KW = {
    'training':    ['train', 'fine-tun', 'fit the model', 'model training'],
    'evaluation':  ['evaluat', 'ndcg', 'mrr', 'offline test', 'benchmark', 'ablation'],
    'deployment':  ['deploy', 'serving', 'production', 'inference', 'rollout'],
    'monitoring':  ['monitor', 'drift', 'alerting', 'observ', 'logging', 'dashboard'],
}
RETRIEVAL_RANKING_KW = ['retrieval', 'ranking', 'vector', 'faiss', 'pinecone',
                        'weaviate', 'qdrant', 'milvus', 'elasticsearch',
                        'opensearch', 'hybrid search', 'recommendation', 'rerank']
LLM_DEPLOY_KW = ['llm', 'large language model', 'gpt', 'production', 'deploy',
                  'inference', 'fine-tun', 'rag', 'serving']
ARCH_ONLY_KW = ['architecture', 'architect', 'design review', 'stakeholder',
                'roadmap', 'governance']
CODING_KW = ['implement', 'built', 'developed', 'coded', 'shipped', 'wrote',
             'engineered', 'optimi']
AI_KEYWORDS = ['embedding', 'retrieval', 'ranking', 'recommendation',
               'search', 'nlp', 'machine learning', 'deep learning',
               'rag', 'vector', 'llm', 'transformer', 'model training', 'fine-tun']
SENIOR_TITLE_RE = re.compile(r'\b(vp|vice president|director|head of|chief|cto|principal)\b', re.I)
JUNIOR_TITLE_RE = re.compile(r'\b(intern|junior|jr\.?|associate|trainee)\b', re.I)


# ============================================================
# REGEX COMPILATION
# ============================================================

def _compile_patterns(phrases):
    compiled = []
    for phrase in phrases:
        esc = re.escape(phrase.strip().lower())
        if phrase.endswith('-') or phrase.endswith('tun'):
            pattern = r'(?<![a-z0-9])' + esc
        else:
            pattern = r'(?<![a-z0-9])' + esc + r'(?![a-z0-9])'
        compiled.append(re.compile(pattern))
    return compiled

def _any_pattern_matches(patterns, text):
    return any(p.search(text) for p in patterns)

MANDATORY_SKILL_PATTERNS   = _compile_patterns(MANDATORY_SKILLS)
OPTIONAL_SKILL_PATTERNS    = _compile_patterns(OPTIONAL_SKILLS)
NEGATIVE_SKILL_PATTERNS    = _compile_patterns(NEGATIVE_SKILLS)
AI_KEYWORD_PATTERNS        = _compile_patterns(AI_KEYWORDS)
RETRIEVAL_RANKING_PATTERNS = _compile_patterns(RETRIEVAL_RANKING_KW)
LLM_DEPLOY_PATTERNS        = _compile_patterns(LLM_DEPLOY_KW)
ARCH_ONLY_PATTERNS         = _compile_patterns(ARCH_ONLY_KW)
CODING_PATTERNS            = _compile_patterns(CODING_KW)
PRODUCTION_PATTERNS        = _compile_patterns(PRODUCTION_KW)
IMPACT_PATTERNS            = _compile_patterns(IMPACT_KW)
OSS_PATTERNS               = _compile_patterns(OSS_KW)
LEADERSHIP_PATTERNS        = _compile_patterns(LEADERSHIP_KW)
LIFECYCLE_PATTERNS         = {k: _compile_patterns(v) for k, v in LIFECYCLE_KW.items()}


# ============================================================
# HELPERS
# ============================================================

def days_since(date_str):
    try:
        return (TODAY - datetime.strptime(date_str, '%Y-%m-%d').date()).days
    except Exception:
        return 9999

def is_good_title(t):
    tl = t.lower()
    return any(re.search(p, tl) for p in GOOD_TITLE_PATTERNS)

def is_bad_title(t):
    tl = t.lower()
    return any(re.search(p, tl) for p in BAD_TITLE_PATTERNS)

def _seniority_level(title):
    tl = title.lower()
    best = 0
    for word, lvl in SENIORITY_ORDER.items():
        if re.search(r'\b' + re.escape(word) + r'\b', tl):
            best = max(best, lvl)
    return best


# ============================================================
# 1. TITLE SCORE
# ============================================================

def title_score(title):
    if not title or is_bad_title(title):
        return 0.0
    t = title.lower()
    buzz_hits = len(re.findall(r'\b(ai|ml|llm|nlp)\b', t))
    if buzz_hits >= 3 and len(t.split()) <= 8:
        return 0.50
    for score, patterns in TITLE_TIERS:
        if any(re.search(p, t) for p in patterns):
            return score
    if is_good_title(title):
        return 0.85
    if any(x in t for x in ['engineer', 'developer', 'scientist', 'analyst']):
        return 0.30
    return 0.15


# ============================================================
# 2. CAREER SCORE
# ============================================================

def career_score(career_history, profile):
    if not career_history:
        return 0.0, False

    total_months = 0
    retrieval_score_acc = 0.0
    llm_deploy_acc = 0.0
    ai_months_weighted = 0.0
    product_months_w = 0.0
    consulting_months_w = 0.0
    has_production_ai = False
    architecture_only_months = 0
    short_stints = 0
    lifecycle_stages_seen = set()
    has_impact_signals = False
    has_leadership = False
    has_oss = False
    ai_product_company_bonus = 0.0

    seniority_progression = []
    seniority_regression_flag = False

    for job in career_history:
        dur = job.get('duration_months', 0)
        total_months += dur
        company = job.get('company', '').lower()
        industry = job.get('industry', '').lower()
        desc = job.get('description', '').lower()
        title_j = job.get('title', '').lower()
        start_str = job.get('start_date', '')

        end_str = job.get('end_date') or ''
        is_current_role = job.get('is_current', False) or not end_str
        try:
            anchor_d = TODAY if is_current_role else datetime.strptime(end_str, '%Y-%m-%d').date()
            years_ago = (TODAY - anchor_d).days / 365.0
        except Exception:
            years_ago = 5.0
        recency_w = math.exp(-0.15 * years_ago)

        is_consulting_company = any(cf in company for cf in CONSULTING_FIRMS)
        is_consulting_industry = any(ci in industry for ci in CONSULTING_INDUSTRIES)
        has_product_work_in_desc = any(kw in desc for kw in
            ['product', 'platform', 'internal', 'end-to-end', 'shipped', 'own'])
        is_consulting = (is_consulting_company or is_consulting_industry) and not has_product_work_in_desc

        w_dur = dur * recency_w

        if is_consulting:
            consulting_months_w += w_dur
        else:
            product_months_w += w_dur

        if dur > 0 and dur < 6:
            short_stints += 1

        has_prod = _any_pattern_matches(PRODUCTION_PATTERNS, desc)
        has_ai   = _any_pattern_matches(AI_KEYWORD_PATTERNS, desc)
        if has_ai:
            ai_months_weighted += w_dur
            if has_prod:
                has_production_ai = True

        if _any_pattern_matches(RETRIEVAL_RANKING_PATTERNS, desc):
            retrieval_score_acc += w_dur * (1.2 if has_prod else 1.0)

        if _any_pattern_matches(LLM_DEPLOY_PATTERNS, desc):
            llm_deploy_acc += w_dur

        if _any_pattern_matches(ARCH_ONLY_PATTERNS, desc) and not _any_pattern_matches(CODING_PATTERNS, desc):
            architecture_only_months += dur

        for stage, patterns in LIFECYCLE_PATTERNS.items():
            if _any_pattern_matches(patterns, desc):
                lifecycle_stages_seen.add(stage)

        if _any_pattern_matches(IMPACT_PATTERNS, desc):
            has_impact_signals = True

        if _any_pattern_matches(LEADERSHIP_PATTERNS, desc):
            has_leadership = True

        if _any_pattern_matches(OSS_PATTERNS, desc):
            has_oss = True

        if any(apc in company for apc in STRONG_AI_COMPANIES) and has_ai:
            ai_product_company_bonus = min(ai_product_company_bonus + 0.02, 0.06)

        seniority_progression.append((start_str, _seniority_level(job.get('title', ''))))

    total_w = product_months_w + consulting_months_w
    if total_w == 0:
        total_w = max(total_months, 1)

    prod_ratio = product_months_w / total_w
    score = 0.20 * prod_ratio

    ai_yrs_w = ai_months_weighted / 12
    if ai_yrs_w >= 4:
        score += 0.18
    elif ai_yrs_w >= 2:
        score += 0.12
    elif ai_yrs_w >= 1:
        score += 0.06
    elif has_production_ai:
        score += 0.03

    rr_norm = retrieval_score_acc / max(total_months, 1)
    score += 0.22 * min(rr_norm / 0.5, 1.0)

    llm_norm = llm_deploy_acc / max(total_months, 1)
    score += 0.10 * min(llm_norm / 0.4, 1.0)

    if has_production_ai:
        score += 0.05

    stages_covered = len(lifecycle_stages_seen)
    score += 0.05 * (stages_covered / 4)

    if has_impact_signals:
        score += 0.03

    if has_leadership:
        score += 0.02

    score += ai_product_company_bonus

    seniority_progression.sort(key=lambda x: x[0])
    levels = [lvl for _, lvl in seniority_progression]
    if len(levels) >= 2:
        regressions = sum(1 for i in range(1, len(levels)) if levels[i] < levels[i-1] - 1)
        if regressions >= 2:
            seniority_regression_flag = True
            score *= 0.85

    if short_stints >= 3:
        score *= 0.75
    elif short_stints >= 2:
        score *= 0.90

    if total_months > 0 and architecture_only_months / total_months > 0.5:
        score *= 0.70

    consulting_ratio = consulting_months_w / total_w if total_w > 0 else 0
    if consulting_ratio > 0.95 and total_months > 24:
        score *= 0.30
    elif consulting_ratio > 0.70:
        score *= 0.65

    industries = [j.get('industry', '').lower() for j in career_history]
    if industries and all('research' in ind or 'academia' in ind for ind in industries if ind) \
            and not has_production_ai:
        score *= 0.60

    rr_yrs = retrieval_score_acc / 12
    has_strong_ai_career = (rr_yrs >= 1.5) or (ai_yrs_w >= 3) or (llm_norm * total_months / 12 >= 1.5)

    return min(score, 1.0), has_strong_ai_career


# ============================================================
# 3. SKILLS SCORE
# ============================================================

def skills_score(skills, assessment_scores=None, career_history=None):
    if not skills:
        return 0.0
    assessment_scores = assessment_scores or {}
    career_history = career_history or []

    career_corpus = ' '.join(
        j.get('description', '') + ' ' + j.get('title', '')
        for j in career_history
    ).lower()

    total_score = 0.0
    negative_score = 0.0
    mandatory_hits = 0
    optional_hits = 0
    zero_dur_expert = 0
    shallow_skills = 0
    seen_concepts = set()
    skill_names_lower = set()

    for sk in skills:
        name_raw = sk['name']
        name_lower = name_raw.lower()
        prof = sk.get('proficiency', 'beginner')
        endorsements = sk.get('endorsements', 0)
        duration = sk.get('duration_months', 0)
        skill_names_lower.add(name_lower)

        if _any_pattern_matches(NEGATIVE_SKILL_PATTERNS, name_lower):
            negative_score += 0.08
            continue

        is_mandatory = _any_pattern_matches(MANDATORY_SKILL_PATTERNS, name_lower)
        is_optional = (not is_mandatory) and _any_pattern_matches(OPTIONAL_SKILL_PATTERNS, name_lower)

        if not (is_mandatory or is_optional):
            if prof == 'beginner' and duration <= 3 and endorsements == 0:
                shallow_skills += 1
            continue

        concept_key = re.sub(r'[^a-z0-9]', '', name_lower)
        if concept_key in seen_concepts:
            continue
        seen_concepts.add(concept_key)

        if is_mandatory:
            mandatory_hits += 1
        else:
            optional_hits += 1

        tier_weight = 1.0 if is_mandatory else 0.55
        prof_weight = {'beginner': 0.20, 'intermediate': 0.50, 'advanced': 0.80, 'expert': 1.00}.get(prof, 0.30)
        endorse_factor = min(endorsements / 20, 1.0)
        duration_factor = min(duration / 36, 1.0)

        assess_factor = 0.50
        for k, v in assessment_scores.items():
            if k.lower() in name_lower or name_lower in k.lower():
                assess_factor = max(0.0, min(v / 100, 1.0))
                break

        career_cross_val = 0.0
        if re.search(r'(?<![a-z0-9])' + re.escape(name_lower.split()[0]) + r'(?![a-z0-9])', career_corpus):
            career_cross_val = 0.20

        trust = 0.20 + 0.20 * endorse_factor + 0.20 * duration_factor \
               + 0.20 * assess_factor + 0.20 * career_cross_val

        if prof in ('advanced', 'expert') and duration == 0:
            zero_dur_expert += 1

        total_score += prof_weight * trust * tier_weight * 0.15

    if len(skills) >= 25 and mandatory_hits > 0 and mandatory_hits / len(skills) < 0.20:
        total_score *= 0.60

    if optional_hits >= 5 and mandatory_hits <= 1:
        total_score *= 0.75

    if zero_dur_expert >= 4:
        total_score *= 0.50

    if shallow_skills >= 10:
        total_score *= 0.80

    synergy_bonus = 0.0
    for group in SYNERGY_GROUPS:
        matched = sum(1 for concept in group
                      if any(concept in sn for sn in skill_names_lower))
        if matched >= 4:
            synergy_bonus = max(synergy_bonus, 0.08)
        elif matched >= 3:
            synergy_bonus = max(synergy_bonus, 0.04)
    total_score += synergy_bonus

    return max(0.0, min(total_score - negative_score * 0.05, 1.0))


# ============================================================
# 4. EXPERIENCE SCORE
# ============================================================

def experience_score(yoe):
    if yoe <= 0:
        return 0.05
    peak_lo, peak_hi = 6.0, 8.0
    if peak_lo <= yoe <= peak_hi:
        return 1.0
    if yoe < peak_lo:
        dist = peak_lo - yoe
        sigma = 2.2
    else:
        dist = yoe - peak_hi
        sigma = 4.5
    return max(0.10, 0.12 + 0.88 * math.exp(-(dist ** 2) / (2 * sigma ** 2)))


# ============================================================
# 5. LOCATION SCORE
# ============================================================

def location_score(location, country, willing_to_relocate, preferred_work_mode):
    loc_lower = (location + ' ' + country).lower()
    in_top = any(c in loc_lower for c in PREFERRED_LOCATIONS_TOP)
    in_pref = any(c in loc_lower for c in PREFERRED_LOCATIONS)
    in_india = country.lower() in ('india', 'in')
    if in_top:
        return 1.0
    if in_pref:
        return 0.90
    if in_india:
        return 0.75
    if willing_to_relocate:
        return 0.40
    if preferred_work_mode in ('remote', 'flexible'):
        return 0.50
    return 0.10


# ============================================================
# 6. NOTICE SCORE
# ============================================================

def notice_score(notice_days):
    if notice_days <= 0:
        return 1.0
    return max(0.15, math.exp(-notice_days / 60.0))


# ============================================================
# 7. BEHAVIORAL SCORE
# ============================================================

def behavioral_score(signals, career_history=None):
    career_history = career_history or []
    adj = 1.0

    last_active_days = days_since(signals.get('last_active_date', '2020-01-01'))
    if last_active_days <= 7:
        adj *= 1.12
    elif last_active_days <= 30:
        adj *= 1.07
    elif last_active_days <= 90:
        adj *= 1.00
    elif last_active_days <= 180:
        adj *= 0.92
    else:
        adj *= 0.78

    if signals.get('open_to_work_flag'):
        adj *= 1.07

    rr = signals.get('recruiter_response_rate', 0.0)
    adj *= (0.90 + 0.20 * rr)

    rt = signals.get('avg_response_time_hours', 100)
    if rt <= 4:
        adj *= 1.04
    elif rt <= 24:
        adj *= 1.02
    elif rt > 72:
        adj *= 0.97

    icr = signals.get('interview_completion_rate', 0.0)
    adj *= (0.95 + 0.10 * icr)

    gh = signals.get('github_activity_score', -1)
    oss_in_career = any(
        _any_pattern_matches(OSS_PATTERNS, j.get('description', '').lower())
        for j in career_history
    )
    if gh >= 0:
        base_gh = 0.95 + 0.12 * (gh / 100)
        if oss_in_career:
            base_gh = min(base_gh * 1.04, 1.12)
        adj *= base_gh

    if signals.get('verified_email'):
        adj *= 1.01
    if signals.get('verified_phone'):
        adj *= 1.01
    if signals.get('linkedin_connected'):
        adj *= 1.005

    assessments = signals.get('skill_assessment_scores', {})
    relevant = [v for k, v in assessments.items()
                if any(kw in k.lower() for kw in ['python', 'ml', 'nlp', 'ai', 'search', 'retrieval'])]
    if relevant:
        adj *= (0.97 + 0.06 * (sum(relevant) / len(relevant) / 100))

    pc_raw = signals.get('profile_completeness_score', 50)
    pc = max(0.0, min(pc_raw / 100, 1.0))
    adj *= (0.95 + 0.08 * pc)

    saved = signals.get('saved_by_recruiters_30d', 0)
    adj *= (1.0 + 0.04 * min(saved / 8.0, 1.0))

    oar = signals.get('offer_acceptance_rate', -1)
    if oar >= 0:
        adj *= (0.97 + 0.05 * oar)

    return max(0.65, min(adj, 1.25))


# ============================================================
# 8. EDUCATION SCORE
# ============================================================

def education_score(education):
    if not education:
        return 0.30
    tier_order = {'tier_1': 4, 'tier_2': 3, 'tier_3': 2, 'tier_4': 1, 'unknown': 0}
    tier_scores = {'tier_1': 1.0, 'tier_2': 0.70, 'tier_3': 0.45, 'tier_4': 0.25, 'unknown': 0.30}
    relevant_fields = ['computer science', 'software', 'electrical', 'electronics',
                       'statistics', 'mathematics', 'data', 'information', 'machine learning']
    best_tier = 'unknown'
    field_match = False
    for edu in education:
        tier = edu.get('tier', 'unknown')
        if tier_order.get(tier, 0) > tier_order.get(best_tier, 0):
            best_tier = tier
        if any(f in edu.get('field_of_study', '').lower() for f in relevant_fields):
            field_match = True
    base = tier_scores.get(best_tier, 0.30)
    return base * (1.10 if field_match else 0.90)


# ============================================================
# 9. HONEYPOT DETECTION
# ============================================================

def detect_honeypot(candidate):
    career = candidate.get('career_history', [])
    profile = candidate.get('profile', {})
    skills = candidate.get('skills', [])

    parsed_jobs = []
    for job in career:
        start = job.get('start_date', '')
        try:
            start_d = datetime.strptime(start, '%Y-%m-%d').date()
            dur = job.get('duration_months', 0)
            expected = (TODAY - start_d).days / 30
            if dur > expected + 6:
                return True
            parsed_jobs.append((start_d, dur, job.get('title', '')))
        except Exception:
            pass

    parsed_jobs.sort(key=lambda x: x[0])
    for i in range(len(parsed_jobs) - 1):
        s1, d1, _ = parsed_jobs[i]
        s2, d2, _ = parsed_jobs[i + 1]
        if d1 >= 6 and d2 >= 6 and (s1.toordinal() + d1 * 30 - s2.toordinal()) > 180:
            return True

    if len(parsed_jobs) >= 2:
        titles_chrono = [t for _, _, t in parsed_jobs]
        first_start = parsed_jobs[0][0]
        for s, d, t in parsed_jobs:
            if SENIOR_TITLE_RE.search(t):
                if (s - first_start).days / 30 < 24 and any(JUNIOR_TITLE_RE.search(jt) for jt in titles_chrono):
                    return True

    if sum(1 for sk in skills if sk.get('proficiency') == 'expert' and sk.get('duration_months', 1) == 0) >= 3:
        return True

    total_months = sum(j.get('duration_months', 0) for j in career)
    stated_yoe = profile.get('years_of_experience', 0)
    if total_months / 12 > stated_yoe + 5:
        return True

    if parsed_jobs:
        span_years = (TODAY - min(s for s, _, _ in parsed_jobs)).days / 365
        if span_years > stated_yoe + 5:
            return True

    return False


# ============================================================
# 10. COMPOSITE SCORE
# ============================================================

def score_candidate(candidate):
    p = candidate['profile']
    signals = candidate.get('redrob_signals', {})
    career = candidate.get('career_history', [])

    if detect_honeypot(candidate):
        return 0.0, {'honeypot': True}

    current_title = p.get('current_title', '')
    t_sc = title_score(current_title)
    c_sc, has_strong_ai = career_score(career, p)
    sk_sc = skills_score(
        candidate.get('skills', []),
        signals.get('skill_assessment_scores', {}),
        career,
    )
    yoe_sc = experience_score(p.get('years_of_experience', 0))
    loc_sc = location_score(
        p.get('location', ''), p.get('country', ''),
        signals.get('willing_to_relocate', False),
        signals.get('preferred_work_mode', 'onsite'),
    )
    not_sc = notice_score(signals.get('notice_period_days', 90))
    beh   = behavioral_score(signals, career)
    edu_sc = education_score(candidate.get('education', []))

    if has_strong_ai and 0.0 < t_sc <= 0.50:
        t_sc = min(t_sc + 0.30, 0.75)

    if t_sc == 0.0 and c_sc < 0.20 and sk_sc < 0.10:
        return 0.01, {'disqualified': True}

    yoe_val = p.get('years_of_experience', 0)
    seniority_factor = seniority_plausibility_factor(current_title, yoe_val)

    composite = (
        0.22 * t_sc  +
        0.27 * c_sc  +
        0.25 * sk_sc +
        0.10 * yoe_sc +
        0.08 * loc_sc +
        0.05 * not_sc +
        0.03 * edu_sc
    )
    composite *= beh
    composite *= seniority_factor

    components = {
        'title': round(t_sc, 4),
        'career': round(c_sc, 4),
        'skills': round(sk_sc, 4),
        'experience': round(yoe_sc, 4),
        'location': round(loc_sc, 4),
        'notice': round(not_sc, 4),
        'education': round(edu_sc, 4),
        'behavioral_mult': round(beh, 4),
        'seniority_plausibility': round(seniority_factor, 4),
        'composite_pre_norm': round(composite, 4),
    }
    return composite, components


# ============================================================
# 11. SCORE NORMALIZATION
# ============================================================

def normalize_scores(raw_scores):
    if not raw_scores:
        return raw_scores
    n = len(raw_scores)
    mu = sum(raw_scores) / n
    var = sum((x - mu) ** 2 for x in raw_scores) / max(n - 1, 1)
    sigma = math.sqrt(var) if var > 0 else 1.0

    def sigmoid(x):
        return 1.0 / (1.0 + math.exp(-x))

    return [sigmoid((x - mu) / sigma) for x in raw_scores]


# ============================================================
# 12. REASONING GENERATOR
# ============================================================

def generate_reasoning(candidate, components):
    p = candidate['profile']
    signals = candidate.get('redrob_signals', {})
    career = candidate.get('career_history', [])
    skills = candidate.get('skills', [])

    title = p.get('current_title', 'Unknown')
    yoe = p.get('years_of_experience', 0)
    location = p.get('location', '')
    country = p.get('country', '')
    last_active_days = days_since(signals.get('last_active_date', '2020-01-01'))
    open_to_work = signals.get('open_to_work_flag', False)
    notice = signals.get('notice_period_days', 90)
    rr = signals.get('recruiter_response_rate', 0)
    gh = signals.get('github_activity_score', -1)

    top_skills = sorted(
        [sk for sk in skills if
            _any_pattern_matches(MANDATORY_SKILL_PATTERNS, sk['name'].lower()) or
            _any_pattern_matches(OPTIONAL_SKILL_PATTERNS,  sk['name'].lower())],
        key=lambda x: x.get('endorsements', 0), reverse=True
    )
    skill_str = ', '.join(s['name'] for s in top_skills[:2]) if top_skills else None

    retrieval_jobs = [j for j in career if
        _any_pattern_matches(RETRIEVAL_RANKING_PATTERNS, j.get('description', '').lower())]
    ai_jobs = [j for j in career if
        _any_pattern_matches(AI_KEYWORD_PATTERNS, j.get('description', '').lower())]
    product_jobs = [j for j in career if not any(cf in j.get('company', '').lower() for cf in CONSULTING_FIRMS)]

    parts = []
    if retrieval_jobs:
        j = retrieval_jobs[0]
        dur_yrs = min(round(j.get('duration_months', 12) / 12, 1), yoe)
        parts.append(f"{title} with {yoe:.1f}y exp, including {dur_yrs}y on retrieval/ranking/vector-search at {j['company']}")
    elif ai_jobs:
        j = ai_jobs[0]
        dur_yrs = min(round(j.get('duration_months', 12) / 12, 1), yoe)
        parts.append(f"{title} with {yoe:.1f}y exp, including {dur_yrs}y building production AI at {j['company']}")
    elif product_jobs:
        parts.append(f"{title} ({yoe:.1f}y exp) with product-company experience at {product_jobs[0]['company']}")
    else:
        parts.append(f"{title} with {yoe:.1f}y of experience")

    if skill_str:
        parts[0] += f"; skilled in {skill_str}"

    loc_str = f"{location}, {country}" if country and country.lower() != 'india' else location
    avail = []
    if open_to_work:
        avail.append("open to work")
    if last_active_days <= 30:
        avail.append("recently active")
    if notice <= 30:
        avail.append(f"{notice}d notice")
    if rr >= 0.70:
        avail.append("high response rate")
    if gh >= 60:
        avail.append(f"GitHub {gh:.0f}/100")

    concerns = []
    if last_active_days > 90:
        concerns.append(f"inactive {last_active_days}d")
    if rr < 0.20 and last_active_days > 30:
        concerns.append(f"low response ({rr:.0%})")
    if notice > 60:
        concerns.append(f"{notice}d notice")
    if country.lower() not in ('india', 'in') and not signals.get('willing_to_relocate'):
        concerns.append(f"based abroad ({country})")

    comp_str = (f"[scores: career={components.get('career', 0):.2f} "
                f"skills={components.get('skills', 0):.2f} "
                f"beh×={components.get('behavioral_mult', 1):.2f}]")

    s2 = f"Based in {loc_str}"
    if avail:
        s2 += "; " + ", ".join(avail)
    if concerns:
        s2 += "; concerns: " + ", ".join(concerns)
    s2 += ". " + comp_str

    return f"{parts[0]}. {s2}"


# ============================================================
# 13. MAIN
# ============================================================

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Redrob Hackathon Ranker v2")
    parser.add_argument("--candidates", default=None,
                        help="Path to candidates.jsonl (default: same folder as script)")
    parser.add_argument("--out", default=None,
                        help="Output XLSX path (default: submission.xlsx next to script)")
    args = parser.parse_args()

    BASE_DIR = Path(__file__).parent
    input_path = Path(args.candidates) if args.candidates else BASE_DIR / "candidates.jsonl"
    output_path = Path(args.out) if args.out else BASE_DIR / "submission.xlsx"

    print("Loading candidates...", flush=True)
    candidates = []
    with open(input_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                candidates.append(json.loads(line))
    print(f"Loaded {len(candidates)} candidates. Scoring...", flush=True)

    raw_scores = []
    all_components = []

    for i, c in enumerate(candidates):
        sc, comp = score_candidate(c)
        raw_scores.append(sc)
        all_components.append(comp)
        if (i + 1) % 10000 == 0:
            print(f"  {i+1}/{len(candidates)} scored...", flush=True)

    print("Sorting candidates...", flush=True)
    indexed = sorted(
        enumerate(raw_scores),
        key=lambda x: (-x[1], candidates[x[0]]['candidate_id'])
    )
    top100_idx = [idx for idx, _ in indexed[:100]]
    top100_raw = [raw_scores[i] for i in top100_idx]

    top100_norm = normalize_scores(top100_raw)

    print("Writing submission.xlsx...", flush=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    headers = ['candidate_id', 'rank', 'score', 'reasoning']
    ws.append(headers)
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='2F5496')
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    body_font = Font(name='Arial')
    for rank, (cand_idx, norm_sc) in enumerate(zip(top100_idx, top100_norm), 1):
        c = candidates[cand_idx]
        comp = all_components[cand_idx]
        reasoning = generate_reasoning(c, comp)
        row = [c['candidate_id'], rank, round(norm_sc, 6), reasoning]
        ws.append(row)
        r = ws.max_row
        for col_idx in range(1, 5):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = body_font
            if col_idx in (2, 3):
                cell.alignment = Alignment(horizontal='center')
            if col_idx == 4:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions[get_column_letter(1)].width = 16
    ws.column_dimensions[get_column_letter(2)].width = 8
    ws.column_dimensions[get_column_letter(3)].width = 10
    ws.column_dimensions[get_column_letter(4)].width = 110

    wb.save(output_path)

    print(f"\nDone! Written to {output_path}")
    print("\nTop 10 candidates:")
    for rank, cand_idx in enumerate(top100_idx[:10], 1):
        c = candidates[cand_idx]
        p = c['profile']
        comp = all_components[cand_idx]
        print(f"  {rank:2d}. {c['candidate_id']} | {p['current_title'][:40]} | "
              f"{p['years_of_experience']}y | {p['location']} | "
              f"raw={raw_scores[cand_idx]:.4f} "
              f"career={comp.get('career',0):.2f} "
              f"skills={comp.get('skills',0):.2f} "
              f"beh={comp.get('behavioral_mult',1):.2f}")


if __name__ == '__main__':
    main()
